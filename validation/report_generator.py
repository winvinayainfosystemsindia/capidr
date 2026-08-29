"""Excel report builder — produces a styled .xlsx validation report."""

from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter

from .models import CheckStatus, ValidationReport


# ---------------------------------------------------------------------------
# Style constants
# ---------------------------------------------------------------------------

_HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
_HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")

_PASS_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
_PASS_FONT = Font(name="Calibri", size=10, color="006100")

_FAIL_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
_FAIL_FONT = Font(name="Calibri", size=10, color="9C0006")

_NA_FILL = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
_NA_FONT = Font(name="Calibri", size=10, color="595959")

_BODY_FONT = Font(name="Calibri", size=10)
_BODY_ALIGNMENT = Alignment(wrap_text=True, vertical="top")

_THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

_SUMMARY_LABEL_FONT = Font(name="Calibri", size=11, bold=True)
_SUMMARY_VALUE_FONT = Font(name="Calibri", size=11)
_SUMMARY_TITLE_FONT = Font(name="Calibri", size=16, bold=True, color="1F4E79")

# Preferred column widths (approximate character widths)
_COLUMN_WIDTHS = {
    "A": 28,   # Section
    "B": 30,   # Checklist Item
    "C": 45,   # Description
    "D": 35,   # WCAG 2.2 Criteria
    "E": 10,   # Status
    "F": 45,   # Failure Reason
    "G": 40,   # Location
    "H": 40,   # Expected Value
    "I": 40,   # Actual Value
}


def generate_excel_report(report: ValidationReport, output_path: str) -> str:
    """Generate a styled Excel validation report.

    Creates two sheets:
      1. **Summary** — high-level pass/fail/N/A stats
      2. **Validation Details** — every individual check result

    Args:
        report: The completed ValidationReport.
        output_path: Where to save the .xlsx file.

    Returns:
        The absolute path to the generated file.
    """
    wb = Workbook()

    # -----------------------------------------------------------------------
    # Sheet 1: Summary
    # -----------------------------------------------------------------------
    ws_summary = wb.active
    ws_summary.title = "Summary"
    ws_summary.sheet_properties.tabColor = "1F4E79"

    # Title
    ws_summary.merge_cells("A1:D1")
    title_cell = ws_summary["A1"]
    title_cell.value = "WCAG 2.2 Validation Report"
    title_cell.font = _SUMMARY_TITLE_FONT
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws_summary.row_dimensions[1].height = 35

    # Metadata
    summary_data = [
        ("Document:", Path(report.document_path).name),
        ("Report Generated:", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("", ""),
        ("Total Checks:", report.total_checks),
        ("Passed:", report.pass_count),
        ("Failed:", report.fail_count),
        ("Not Applicable:", report.na_count),
        ("Pass Rate (applicable):", f"{report.pass_rate}%"),
    ]

    for row_offset, (label, value) in enumerate(summary_data, start=3):
        label_cell = ws_summary.cell(row=row_offset, column=1, value=label)
        label_cell.font = _SUMMARY_LABEL_FONT

        value_cell = ws_summary.cell(row=row_offset, column=2, value=value)
        value_cell.font = _SUMMARY_VALUE_FONT

        # Color-code pass rate
        if label == "Pass Rate (applicable):":
            rate = report.pass_rate
            if rate >= 90:
                value_cell.fill = _PASS_FILL
                value_cell.font = Font(name="Calibri", size=11, bold=True, color="006100")
            elif rate >= 70:
                value_cell.fill = PatternFill(
                    start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"
                )
                value_cell.font = Font(name="Calibri", size=11, bold=True, color="9C6500")
            else:
                value_cell.fill = _FAIL_FILL
                value_cell.font = Font(name="Calibri", size=11, bold=True, color="9C0006")

    # Summary column widths
    ws_summary.column_dimensions["A"].width = 25
    ws_summary.column_dimensions["B"].width = 40

    # Section breakdown table
    section_row = len(summary_data) + 5
    ws_summary.cell(row=section_row, column=1, value="Section Breakdown").font = (
        Font(name="Calibri", size=13, bold=True, color="1F4E79")
    )
    section_row += 1

    # Headers
    for col_idx, header in enumerate(["Section", "Pass", "Fail", "N/A", "Total"], 1):
        cell = ws_summary.cell(row=section_row, column=col_idx, value=header)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.border = _THIN_BORDER
        cell.alignment = Alignment(horizontal="center")

    # Aggregate by section
    section_stats = {}
    for result in report.results:
        if result.section not in section_stats:
            section_stats[result.section] = {"PASS": 0, "FAIL": 0, "N/A": 0}
        section_stats[result.section][result.status.value] += 1

    for s_offset, (section_name, stats) in enumerate(section_stats.items(), 1):
        row = section_row + s_offset
        total = stats["PASS"] + stats["FAIL"] + stats["N/A"]
        for col_idx, value in enumerate(
            [section_name, stats["PASS"], stats["FAIL"], stats["N/A"], total], 1
        ):
            cell = ws_summary.cell(row=row, column=col_idx, value=value)
            cell.font = _BODY_FONT
            cell.border = _THIN_BORDER
            if col_idx > 1:
                cell.alignment = Alignment(horizontal="center")

    # Adjust section breakdown column widths
    ws_summary.column_dimensions["C"].width = 10
    ws_summary.column_dimensions["D"].width = 10
    ws_summary.column_dimensions["E"].width = 10

    # -----------------------------------------------------------------------
    # Sheet 2: Validation Details
    # -----------------------------------------------------------------------
    ws_detail = wb.create_sheet("Validation Details")
    ws_detail.sheet_properties.tabColor = "2E75B6"

    # Write headers
    headers = report.headers
    for col_idx, header in enumerate(headers, 1):
        cell = ws_detail.cell(row=1, column=col_idx, value=header)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.border = _THIN_BORDER
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )

    ws_detail.row_dimensions[1].height = 30

    # Write data rows
    for row_idx, result in enumerate(report.results, start=2):
        row_data = result.to_row()
        for col_idx, value in enumerate(row_data, 1):
            cell = ws_detail.cell(row=row_idx, column=col_idx, value=value)
            cell.font = _BODY_FONT
            cell.alignment = _BODY_ALIGNMENT
            cell.border = _THIN_BORDER

        # Apply status-based formatting to the Status column (E) and full row
        status_cell = ws_detail.cell(row=row_idx, column=5)
        if result.status == CheckStatus.PASS:
            status_cell.fill = _PASS_FILL
            status_cell.font = Font(name="Calibri", size=10, bold=True, color="006100")
        elif result.status == CheckStatus.FAIL:
            status_cell.fill = _FAIL_FILL
            status_cell.font = Font(name="Calibri", size=10, bold=True, color="9C0006")
        elif result.status == CheckStatus.NOT_APPLICABLE:
            status_cell.fill = _NA_FILL
            status_cell.font = Font(name="Calibri", size=10, bold=True, color="595959")

        # Light row fill for FAIL rows
        if result.status == CheckStatus.FAIL:
            light_fail = PatternFill(
                start_color="FFF2F2", end_color="FFF2F2", fill_type="solid"
            )
            for col_idx in range(1, len(row_data) + 1):
                if col_idx != 5:  # skip status column, already styled
                    ws_detail.cell(row=row_idx, column=col_idx).fill = light_fail

    # Set column widths
    for col_letter, width in _COLUMN_WIDTHS.items():
        ws_detail.column_dimensions[col_letter].width = width

    # Freeze top row + auto-filter
    ws_detail.freeze_panes = "A2"
    ws_detail.auto_filter.ref = (
        f"A1:{get_column_letter(len(headers))}{len(report.results) + 1}"
    )

    # -----------------------------------------------------------------------
    # Save
    # -----------------------------------------------------------------------
    wb.save(output_path)
    return str(Path(output_path).resolve())
